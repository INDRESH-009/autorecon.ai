"""Common openpyxl styling helpers shared across the four reports."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F2937")  # slate-800
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
SUBHEAD_FILL = PatternFill("solid", start_color="E5E7EB")  # gray-200
SUBHEAD_FONT = Font(name=FONT, bold=True, color="111827", size=11)
BODY_FONT = Font(name=FONT, size=10)
BOLD_BODY_FONT = Font(name=FONT, size=10, bold=True)

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILLS = {
    "ok_to_pay":         PatternFill("solid", start_color="DCFCE7"),  # green-100
    "not_due":           PatternFill("solid", start_color="DBEAFE"),  # blue-100
    "pending_in_bt":     PatternFill("solid", start_color="FEF3C7"),  # amber-100
    "to_be_booked":      PatternFill("solid", start_color="FEE2E2"),  # red-100
    "missing_in_vendor": PatternFill("solid", start_color="FCE7F3"),  # pink-100
    "amount_dispute":    PatternFill("solid", start_color="F3E8FF"),  # purple-100
    "paid":              PatternFill("solid", start_color="E5E7EB"),  # gray-200
}


def new_workbook():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def write_header_row(sheet, row_idx, columns, widths=None):
    for i, col in enumerate(columns, 1):
        cell = sheet.cell(row=row_idx, column=i, value=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER
    sheet.row_dimensions[row_idx].height = 28
    if widths:
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = w


def write_body_cell(cell, value, status=None, bold=False, fmt=None):
    cell.value = value
    cell.font = BOLD_BODY_FONT if bold else BODY_FONT
    cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=False)
    cell.border = BORDER
    if status and status in STATUS_FILLS:
        cell.fill = STATUS_FILLS[status]
    if fmt:
        cell.number_format = fmt


def autosize(sheet, max_col):
    """Light autosize — measures longest cell value per column, capped."""
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in sheet[letter]:
            v = cell.value
            if v is None:
                continue
            ln = len(str(v))
            if ln > max_len:
                max_len = ln
        sheet.column_dimensions[letter].width = min(max_len + 2, 45)
