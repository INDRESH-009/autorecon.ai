"""XLSX extractor. Uses openpyxl directly so we can pick header rows and stay
generic across messy real-world layouts."""
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.extraction.coerce import dedupe_headers, jsonable
from app.extraction.types import ExtractedTable


def extract(
    path: str | Path,
    sheet_name: str | None = None,
    header_row: int = 0,
) -> ExtractedTable:
    """Extract a single sheet from an .xlsx file.

    header_row is 0-indexed (row 0 = first row in the sheet).
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    all_rows = list(sheet.iter_rows(values_only=True))
    if header_row >= len(all_rows):
        return ExtractedTable(headers=[], rows=[], sheet=sheet.title)

    raw_headers = list(all_rows[header_row])
    headers = dedupe_headers(raw_headers)

    rows: list = []
    for raw in all_rows[header_row + 1 :]:
        # Skip rows that are entirely empty.
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in raw):
            continue
        row = {}
        for h, cell in zip(headers, raw):
            row[h] = jsonable(cell)
        rows.append(row)
    return ExtractedTable(headers=headers, rows=rows, sheet=sheet.title)


def list_sheets(path: str | Path) -> list:
    wb = load_workbook(path, read_only=True, data_only=True)
    return list(wb.sheetnames)
